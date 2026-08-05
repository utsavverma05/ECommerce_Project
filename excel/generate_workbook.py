
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, CellIsRule, FormulaRule
)
from openpyxl.utils.dataframe import dataframe_to_rows

from analysis.config import OUTPUT_DIR
from analysis.db_loader import get_conn

EXCEL_DIR = Path(__file__).resolve().parent
OUT_PATH  = EXCEL_DIR / "ops_case_study.xlsx"

# -- Style tokens --------------------------------------------------------------
CLR = {
    "header_bg":  "1A1A2E",
    "header_fg":  "FFFFFF",
    "accent":     "2E86AB",
    "danger":     "E84855",
    "warning":    "F4A261",
    "success":    "2D9B72",
    "row_alt":    "F2F6FA",
    "row_main":   "FFFFFF",
    "border":     "D0D7DE",
}

HEADER_FONT  = Font(name="Calibri", bold=True, color=CLR["header_fg"], size=10)
HEADER_FILL  = PatternFill("solid", fgColor=CLR["header_bg"])
TITLE_FONT   = Font(name="Calibri", bold=True, size=13, color=CLR["accent"])
BODY_FONT    = Font(name="Calibri", size=9)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN         = Side(style="thin", color=CLR["border"])
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_df(ws, df, start_row=3, col_widths=None):
    """Write a DataFrame to a worksheet starting at start_row."""
    # Header
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c, value=col)
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    # Data rows
    for r, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = BODY_FONT
            cell.alignment = LEFT
            cell.border    = BORDER
            if (r - start_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=CLR["row_alt"])

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _title(ws, text, subtitle=""):
    ws["A1"] = text
    ws["A1"].font      = TITLE_FONT
    ws["A1"].alignment = LEFT
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font      = Font(name="Calibri", size=9, italic=True, color="666666")
        ws["A2"].alignment = LEFT


def _rag_fill(ws, col_letter, start, end, thresholds, fmt=None):
    """Apply green/amber/red conditional formatting to a column."""
    rng = f"{col_letter}{start}:{col_letter}{end}"
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="greaterThan", formula=[str(thresholds[1])],
        fill=PatternFill("solid", fgColor=CLR["success"])
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="between", formula=[str(thresholds[0]), str(thresholds[1])],
        fill=PatternFill("solid", fgColor=CLR["warning"])
    ))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator="lessThan", formula=[str(thresholds[0])],
        fill=PatternFill("solid", fgColor=CLR["danger"])
    ))


def _data_bar(ws, col_letter, start, end, color=None):
    color = color or CLR["accent"]
    rng = f"{col_letter}{start}:{col_letter}{end}"
    ws.conditional_formatting.add(rng, DataBarRule(
        start_type="min", start_value=0,
        end_type="max", end_value=100,
        color=color
    ))


# -- Load data -----------------------------------------------------------------
def _load():
    conn = get_conn()

    otd = pd.read_sql("""
        SELECT s.seller_state AS Node,
               STRFTIME('%Y-%m', o.order_purchase_timestamp) AS Month,
               COUNT(DISTINCT o.order_id) AS Orders,
               ROUND(100.0*SUM(CASE WHEN o.order_delivered_customer_date<=o.order_estimated_delivery_date THEN 1 ELSE 0 END)/COUNT(DISTINCT o.order_id),1) AS [OTD %],
               ROUND(AVG(JULIANDAY(o.order_delivered_customer_date)-JULIANDAY(o.order_estimated_delivery_date)),2) AS [Avg Delay (d)]
        FROM orders o JOIN order_items oi ON o.order_id=oi.order_id JOIN sellers s ON oi.seller_id=s.seller_id
        WHERE o.order_status='delivered' AND o.order_delivered_customer_date IS NOT NULL
        GROUP BY 1,2 ORDER BY 1,2""", conn)

    rto = pd.read_sql("""
        SELECT s.seller_state AS Node,
               COUNT(DISTINCT o.order_id) AS [Total Orders],
               SUM(CASE WHEN o.order_status IN ('cancelled','unavailable') OR o.order_delivered_customer_date IS NULL THEN 1 ELSE 0 END) AS [RTO Orders],
               ROUND(100.0*SUM(CASE WHEN o.order_status IN ('cancelled','unavailable') OR o.order_delivered_customer_date IS NULL THEN 1 ELSE 0 END)/COUNT(DISTINCT o.order_id),1) AS [RTO %],
               ROUND(SUM(CASE WHEN o.order_status IN ('cancelled','unavailable') OR o.order_delivered_customer_date IS NULL THEN oi.price+oi.freight_value ELSE 0 END),0) AS [Revenue at Risk]
        FROM orders o JOIN order_items oi ON o.order_id=oi.order_id JOIN sellers s ON oi.seller_id=s.seller_id
        GROUP BY 1 ORDER BY 4 DESC""", conn)

    bottleneck = pd.read_sql("""
        SELECT s.seller_state AS Node,
               ROUND(AVG(JULIANDAY(o.order_delivered_carrier_date)-JULIANDAY(o.order_approved_at)),2) AS [Processing Days],
               ROUND(AVG(JULIANDAY(o.order_delivered_customer_date)-JULIANDAY(o.order_delivered_carrier_date)),2) AS [Transit Days],
               ROUND(AVG(JULIANDAY(o.order_delivered_customer_date)-JULIANDAY(o.order_approved_at)),2) AS [Total Days]
        FROM orders o JOIN order_items oi ON o.order_id=oi.order_id JOIN sellers s ON oi.seller_id=s.seller_id
        WHERE o.order_status='delivered' AND o.order_delivered_customer_date IS NOT NULL AND o.order_delivered_carrier_date IS NOT NULL
        GROUP BY 1 ORDER BY 4 DESC""", conn)

    abc = pd.read_sql("""
        SELECT oi.seller_id AS [Seller ID], s.seller_state AS State,
               COUNT(DISTINCT o.order_id) AS Orders,
               ROUND(SUM(oi.price+oi.freight_value),0) AS Revenue,
               ROUND(100.0*SUM(CASE WHEN o.order_delivered_customer_date<=o.order_estimated_delivery_date THEN 1 ELSE 0 END)/COUNT(DISTINCT o.order_id),1) AS [OTD %]
        FROM orders o JOIN order_items oi ON o.order_id=oi.order_id JOIN sellers s ON oi.seller_id=s.seller_id
        WHERE o.order_status='delivered' AND o.order_delivered_customer_date IS NOT NULL
        GROUP BY 1,2 HAVING COUNT(DISTINCT o.order_id)>=15 ORDER BY Revenue DESC""", conn)

    # ABC classification
    abc = abc.sort_values("Revenue", ascending=False).reset_index(drop=True)
    abc["Cum Rev %"] = (abc["Revenue"].cumsum() / abc["Revenue"].sum() * 100).round(1)
    abc["Class"] = pd.cut(abc["Cum Rev %"], bins=[0,70,90,100],
                          labels=["A","B","C"], include_lowest=True).astype(str)

    ohs = pd.read_csv(OUTPUT_DIR / "ops_health_score.csv") if (OUTPUT_DIR / "ops_health_score.csv").exists() else pd.DataFrame()

    csat = pd.read_sql("""
        SELECT STRFTIME('%Y-%m', o.order_purchase_timestamp) AS Month,
               ROUND(AVG(r.review_score),2) AS [Avg Review Score],
               ROUND(AVG(JULIANDAY(o.order_delivered_customer_date)-JULIANDAY(o.order_estimated_delivery_date)),2) AS [Avg Delay (d)],
               COUNT(r.review_id) AS Reviews
        FROM orders o JOIN order_reviews r ON o.order_id=r.order_id
        WHERE o.order_status='delivered' AND o.order_delivered_customer_date IS NOT NULL
        GROUP BY 1 ORDER BY 1""", conn)

    trend = pd.read_sql("""
        SELECT STRFTIME('%Y-%m', order_purchase_timestamp) AS Month,
               COUNT(order_id) AS Orders,
               ROUND(100.0*SUM(CASE WHEN order_delivered_customer_date<=order_estimated_delivery_date THEN 1 ELSE 0 END)/COUNT(order_id),1) AS [OTD %],
               ROUND(AVG(JULIANDAY(order_delivered_customer_date)-JULIANDAY(order_estimated_delivery_date)),2) AS [Avg Delay (d)],
               ROUND(AVG(JULIANDAY(order_delivered_carrier_date)-JULIANDAY(order_approved_at)),2) AS [Avg Processing (d)]
        FROM orders WHERE order_status='delivered' AND order_delivered_customer_date IS NOT NULL
        GROUP BY 1 ORDER BY 1""", conn)

    conn.close()
    return otd, rto, bottleneck, abc, ohs, csat, trend


# -- Build workbook ------------------------------------------------------------
def build():
    print("\n[Excel] Building workbook ...")
    otd, rto, bottleneck, abc, ohs, csat, trend = _load()
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # -- Sheet 1: OTD Dashboard ------------------------------------------------
    ws = wb.create_sheet("1_OTD_Dashboard")
    ws.sheet_properties.tabColor = CLR["accent"]
    _title(ws, "OTD Dashboard", "Decision: Executive health check -- which nodes to escalate this month")
    _write_df(ws, otd, col_widths=[6, 9, 9, 10, 14])
    n = len(otd) + 3
    _rag_fill(ws, "D", 4, n, [75, 85])
    _data_bar(ws, "E", 4, n, CLR["danger"])

    # -- Sheet 2: RTO Analysis -------------------------------------------------
    ws2 = wb.create_sheet("2_RTO_Analysis")
    ws2.sheet_properties.tabColor = CLR["danger"]
    _title(ws2, "RTO Analysis", "Decision: Last-mile partner audit; nodes to deprioritize")
    _write_df(ws2, rto, col_widths=[6, 14, 12, 9, 18])
    n2 = len(rto) + 3
    _rag_fill(ws2, "D", 4, n2, [5, 10])   # inverted: <5 good, >10 critical
    _data_bar(ws2, "E", 4, n2, CLR["danger"])

    # -- Sheet 3: Delay Deep Dive ----------------------------------------------
    ws3 = wb.create_sheet("3_Delay_Deep_Dive")
    ws3.sheet_properties.tabColor = CLR["warning"]
    _title(ws3, "Delay Deep Dive", "Decision: Blame attribution -- seller vs carrier bottleneck per node")
    bottleneck["Bottleneck"] = bottleneck.apply(
        lambda r: "Warehouse" if r["Processing Days"] > r["Transit Days"] else "Carrier", axis=1)
    bottleneck["Processing %"] = (bottleneck["Processing Days"] / bottleneck["Total Days"] * 100).round(1)
    _write_df(ws3, bottleneck, col_widths=[6, 18, 14, 12, 14, 14])
    _data_bar(ws3, "B", 4, len(bottleneck) + 3, CLR["accent"])

    # -- Sheet 4: ABC Analysis -------------------------------------------------
    ws4 = wb.create_sheet("4_ABC_Analysis")
    ws4.sheet_properties.tabColor = CLR["success"]
    _title(ws4, "ABC Seller Analysis", "Decision: Priority SLA enforcement -- Class A sellers with low OTD = highest risk")
    _write_df(ws4, abc.head(200), col_widths=[36, 8, 10, 14, 10, 12, 8])
    n4 = min(len(abc), 200) + 3
    _rag_fill(ws4, "E", 4, n4, [75, 85])

    # -- Sheet 5: Revenue at Risk -----------------------------------------------
    ws5 = wb.create_sheet("5_Revenue_At_Risk")
    ws5.sheet_properties.tabColor = CLR["danger"]
    _title(ws5, "Revenue at Risk", "Decision: Finance case for remediation investment")
    risk_df = rto[["Node","Total Orders","RTO Orders","RTO %","Revenue at Risk"]].copy()
    risk_df["Est Annual Loss"] = (risk_df["Revenue at Risk"] * 2).round(0)
    _write_df(ws5, risk_df, col_widths=[6, 14, 12, 9, 18, 18])
    _data_bar(ws5, "E", 4, len(risk_df) + 3, CLR["danger"])

    # -- Sheet 6: Node Risk Matrix ---------------------------------------------
    ws6 = wb.create_sheet("6_Node_Risk_Matrix")
    ws6.sheet_properties.tabColor = CLR["warning"]
    _title(ws6, "Node Risk Matrix", "Decision: Network topology -- nodes x states with worst delivery time")
    if not ohs.empty:
        pivot = ohs.pivot_table(index="node", columns="month", values="ohs", aggfunc="mean").round(1)
        pivot = pivot.reset_index().rename(columns={"node": "Node"})
        _write_df(ws6, pivot, col_widths=[8] + [12] * (len(pivot.columns) - 1))
        n6 = len(pivot) + 3
        for c in range(2, len(pivot.columns) + 1):
            _rag_fill(ws6, get_column_letter(c), 4, n6, [60, 80])
    else:
        ws6["A3"] = "Run analysis/05_ops_health_score.py first to generate OHS data."

    # -- Sheet 7: CSAT Correlation ---------------------------------------------
    ws7 = wb.create_sheet("7_CSAT_Correlation")
    ws7.sheet_properties.tabColor = CLR["success"]
    _title(ws7, "CSAT vs Delay Correlation", "Decision: Quantify CX cost of ops failure")
    _write_df(ws7, csat, col_widths=[9, 18, 16, 10])
    _rag_fill(ws7, "B", 4, len(csat) + 3, [3.5, 4.0])

    # -- Sheet 8: Ops Health Score ---------------------------------------------
    ws8 = wb.create_sheet("8_Ops_Health_Score")
    ws8.sheet_properties.tabColor = CLR["accent"]
    _title(ws8, "Operations Health Score (OHS)",
           "Decision: Monthly scorecard -- trigger SLA review for OHS < 60")
    if not ohs.empty:
        ohs_disp = ohs[["node","month","total_orders","otd_rate","rto_rate",
                         "avg_delay_days","avg_proc_days","ohs","status"]].copy()
        ohs_disp.columns = ["Node","Month","Orders","OTD Rate","RTO Rate",
                             "Avg Delay (d)","Avg Proc (d)","OHS Score","Status"]
        ohs_disp["OTD Rate"] = (ohs_disp["OTD Rate"] * 100).round(1)
        ohs_disp["RTO Rate"] = (ohs_disp["RTO Rate"] * 100).round(1)
        _write_df(ws8, ohs_disp, col_widths=[6,9,9,10,10,13,13,12,10])
        n8 = len(ohs_disp) + 3
        _rag_fill(ws8, "H", 4, n8, [60, 80])
    else:
        ws8["A3"] = "Run analysis/05_ops_health_score.py first."

    wb.save(OUT_PATH)
    print(f"   [OK] Workbook saved -> {OUT_PATH.name}  ({OUT_PATH.stat().st_size // 1024} KB)")
    print(f"     8 sheets | RAG formatting | data bars | freeze panes")


if __name__ == "__main__":
    build()
